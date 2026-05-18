#!/usr/bin/env python3
"""
BancoBanQuito - Generador de Excel
Replica la logica REAL del DataInitializer.java tal como corrio en el servidor:
  - Clientes creados en orden naturalIndex 1..500 (orden de insercion)
  - Cuentas asignadas en ese mismo orden (sin sort por cedula)
  - Primeros 100 en orden de creacion reciben 2 cuentas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Arrays identicos al Java ──────────────────────────────────────────────────
NOMBRES = [
    "Juan","Maria","Carlos","Ana","Luis","Gabriela","Pedro","Daniela",
    "Jorge","Paola","Andres","Camila","Diego","Valeria","Fernando",
    "Sofia","Mateo","Carolina","Ricardo","Isabel","Sebastian","Diana",
    "Esteban","Fernanda","Cristian","Andrea","Mauricio","Karla"
]  # len = 28

APELLIDOS = [
    "Perez","Garcia","Morales","Vera","Cevallos","Mendoza","Castro",
    "Zambrano","Rojas","Sanchez","Ortiz","Torres","Salazar","Guerrero",
    "Navarrete","Paredes","Espinoza","Romero","Alvarez","Delgado",
    "Molina","Quintero","Benitez","Cabrera","Vargas","Acosta"
]  # len = 26

COMPANY_NAMES = [
    "Inversiones Andinas del Ecuador S.A.",
    "Comercializadora Pichincha Cía. Ltda.",
    "Grupo Empresarial Pacífico S.A.",
    "Importadora Continental Cía. Ltda.",
    "Distribuciones Sierra Verde S.A.",
    "Tecnologías Innovación Ecuador S.A.",
    "Construcciones Metropolitanas S.A.",
    "Agroindustrial Cóndor Cía. Ltda.",
    "Transportes Nacionales Unidos S.A.",
    "Soluciones Corporativas Andes S.A.",
    "Industrias Alimenticias del Norte Cía. Ltda.",
    "Grupo Logístico Manabí S.A.",
    "Servicios Financieros Austral Cía. Ltda.",
    "Exportadora Amazónica S.A.",
    "Consultora Empresarial Cuenca Cía. Ltda.",
    "Manufactura Especializada Guayas S.A.",
    "Telecomunicaciones Nacionales S.A.",
    "Inmobiliaria Capital Norte Cía. Ltda.",
    "Seguridad Integral Ecuatoriana S.A.",
    "Farmacéutica Andina Cía. Ltda.",
    "Energía Renovable del Ecuador S.A.",
    "Textiles del Oriente Cía. Ltda.",
    "Automotriz Nacional S.A.",
    "Hotelería y Turismo Galápagos Cía. Ltda.",
    "Alimentos Procesados del Sur S.A.",
    "Ingeniería Civil y Arquitectura Cía. Ltda.",
    "Servicios de Salud Integral S.A.",
    "Tecnología Agropecuaria Nacional Cía. Ltda.",
    "Retail y Comercio Especializado S.A.",
    "Grupo Empresarial Tungurahua Cía. Ltda.",
    "Distribuciones Comerciales Loja S.A.",
    "Petroquímica Ecuatoriana Cía. Ltda.",
    "Ganadería y Producción Agropecuaria S.A.",
    "Centro Comercial Metropolitano Cía. Ltda.",
    "Producción Audiovisual Nacional S.A.",
    "Gestión Ambiental Sostenible Cía. Ltda.",
    "Industria Plástica Especializada S.A.",
    "Operaciones Mineras del Norte Cía. Ltda.",
    "Desarrollo Inmobiliario Moderno S.A.",
    "Servicios Informáticos Avanzados Cía. Ltda.",
    "Exportaciones Marítimas Ecuatorianas S.A.",
    "Procesadora de Alimentos Nativos Cía. Ltda.",
    "Construcciones Viales Nacionales S.A.",
    "Servicios Educativos Superiores Cía. Ltda.",
    "Distribución Eléctrica Nacional S.A.",
    "Laboratorios del Austro Cía. Ltda.",
    "Corporación Manufacturera del Pacífico S.A.",
    "Recursos Naturales Amazónicos Cía. Ltda.",
    "Innovación Biotecnológica Ecuador S.A.",
    "Servicios Portuarios Nacionales Cía. Ltda.",
]  # len = 50

# branchRepository.findAll() → orden de insercion
BRANCHES = [
    {"code": "001", "name": "Sucursal Norte"},
    {"code": "002", "name": "Sucursal Sur"},
    {"code": "003", "name": "Sucursal Centro"},
    {"code": "004", "name": "Sucursal Valles"},
    {"code": "005", "name": "Sucursal Digital"},
]

# ── Algoritmos identicos al Java ──────────────────────────────────────────────
def generate_cedula(index):
    province = (index % 24) + 1
    province_code = f"{province:02d}"
    sequence = 100000 + index
    seq_str = f"{sequence:06d}"[:6]
    base = province_code + seq_str + "0"
    coefficients = [2, 1, 2, 1, 2, 1, 2, 1, 2]
    total = 0
    for i in range(9):
        value = int(base[i]) * coefficients[i]
        if value >= 10:
            value -= 9
        total += value
    verifier = 0 if total % 10 == 0 else 10 - (total % 10)
    return base + str(verifier)

def generate_ruc(index):
    return "179" + f"{index:07d}" + "001"

def account_number(branch_code, seq):
    return branch_code + f"{seq:07d}"

# ── FASE 1: initMassiveCustomers() ────────────────────────────────────────────
# Lista en ORDEN DE CREACION (naturalIndex 1, 2, ..., 500).
# La BD fue cargada SIN el sort por cedula, por eso el orden es de insercion.

naturals_in_order = []   # orden de insercion = orden de naturalIndex
seen_cedulas = set()
natural_index = 1
while len(naturals_in_order) < 500:
    cedula = generate_cedula(natural_index)
    if cedula not in seen_cedulas:
        nombre = NOMBRES[natural_index % len(NOMBRES)]
        ap1    = APELLIDOS[natural_index % len(APELLIDOS)]
        ap2    = APELLIDOS[(natural_index + 9) % len(APELLIDOS)]
        naturals_in_order.append({
            "cedula":        cedula,
            "nombre":        f"{nombre} {ap1} {ap2}",
            "email":         f"{nombre.lower()}.{ap1.lower()}.{ap2.lower()}{natural_index}@banquito.com",
            "tel":           f"09{natural_index:08d}",
            "natural_index": natural_index,
        })
        seen_cedulas.add(cedula)
    natural_index += 1

companies_in_order = []
seen_rucs = set()
corp_index = 1
while len(companies_in_order) < 50:
    ruc = generate_ruc(corp_index)
    if ruc not in seen_rucs:
        companies_in_order.append({
            "ruc":        ruc,
            "nombre":     COMPANY_NAMES[(corp_index - 1) % len(COMPANY_NAMES)],
            "email":      f"contacto.empresa{corp_index}@banquito.com",
            "corp_index": corp_index,
        })
        seen_rucs.add(ruc)
    corp_index += 1

# ── FASE 2: initMassiveAccounts() — orden de insercion, sin sort ──────────────
seq = 1

# Paso A: 1 cuenta AHO para cada natural (orden de creacion)
for i, nat in enumerate(naturals_in_order):
    b = BRANCHES[i % 5]
    nat["cta_ahorros"]   = account_number(b["code"], seq); seq += 1
    nat["suc_ahorros"]   = b["name"]
    nat["cta_corriente"] = None
    nat["suc_corriente"] = None

# Paso B: 1 cuenta CTE para los primeros 100 (en orden de creacion)
for i in range(min(100, len(naturals_in_order))):
    b = BRANCHES[i % 5]
    nat = naturals_in_order[i]
    nat["cta_corriente"] = account_number(b["code"], seq); seq += 1
    nat["suc_corriente"] = b["name"]

# Paso C: 3 cuentas por empresa (CTE→Operativa, NOM→Nomina, AHO→Impuestos)
for i, comp in enumerate(companies_in_order):
    b = BRANCHES[i % 5]
    comp["cta_operativa"]  = account_number(b["code"], seq); seq += 1
    comp["suc_operativa"]  = b["name"]
    comp["cta_nomina"]     = account_number(b["code"], seq); seq += 1
    comp["suc_nomina"]     = b["name"]
    comp["cta_impuestos"]  = account_number(b["code"], seq); seq += 1
    comp["suc_impuestos"]  = b["name"]

# Para el Excel ordenamos por cedula/RUC para facil busqueda
sorted_naturals  = sorted(naturals_in_order,  key=lambda x: x["cedula"])
sorted_companies = sorted(companies_in_order, key=lambda x: x["ruc"])

# ── FASE 3: initDemoData() — sin cuentas del inicializador ────────────────────
DEMO_NATURALS = [
    {"cedula":"1750285577","nombre":"Anahy Herrera Morales",        "email":"anahyherrera09082002@gmail.com",       "tel":"0992832595"},
    {"cedula":"1724356789","nombre":"Carlos Mendoza Rios",          "email":"carlos.mendoza@banquito.fin.ec",       "tel":"0987123456"},
    {"cedula":"1712034567","nombre":"Maria Salazar Vega",           "email":"maria.salazar@banquito.fin.ec",        "tel":"0998234567"},
    {"cedula":"1738901234","nombre":"Luis Ortega Caicedo",          "email":"luis.ortega@banquito.fin.ec",          "tel":"0976345678"},
    {"cedula":"1745678901","nombre":"Gabriela Torres Espinoza",     "email":"gabriela.torres@banquito.fin.ec",      "tel":"0969456789"},
    {"cedula":"1752345678","nombre":"Diego Castro Paredes",         "email":"diego.castro@banquito.fin.ec",         "tel":"0995567890"},
    {"cedula":"1761234567","nombre":"Valeria Guerrero Acosta",      "email":"valeria.guerrero@banquito.fin.ec",     "tel":"0982678901"},
    {"cedula":"1768901234","nombre":"Sebastian Navarrete Ruiz",     "email":"sebastian.navarrete@banquito.fin.ec",  "tel":"0991789012"},
]
DEMO_COMPANIES = [
    {"ruc":"1757158215001","nombre":"TechSolutions Ecuador S.A.",    "email":"info@techsolutions.ec"},
    {"ruc":"1791234567001","nombre":"Importadora Andina Cía. Ltda.", "email":"contacto@importandina.ec"},
    {"ruc":"1791765432001","nombre":"Distribuidora El Comercio S.A.","email":"gerencia@distcomercio.ec"},
]
for d in DEMO_NATURALS:
    d["cta_ahorros"] = "—"; d["suc_ahorros"] = "—"
    d["cta_corriente"] = "—"; d["suc_corriente"] = "—"
for d in DEMO_COMPANIES:
    d["cta_operativa"] = "—"; d["suc_operativa"] = "—"
    d["cta_nomina"] = "—"; d["suc_nomina"] = "—"
    d["cta_impuestos"] = "—"; d["suc_impuestos"] = "—"

# ── Estilos ───────────────────────────────────────────────────────────────────
HDR_FILL_NAT  = PatternFill("solid", fgColor="1a3c6e")
HDR_FILL_CORP = PatternFill("solid", fgColor="0d5c3a")
HDR_FONT      = Font(color="FFFFFF", bold=True, size=10)
DEMO_FILL     = PatternFill("solid", fgColor="fff3cd")
EVEN_N        = PatternFill("solid", fgColor="f4f8ff")
EVEN_C        = PatternFill("solid", fgColor="f0fff4")
ODD_FILL      = PatternFill("solid", fgColor="FFFFFF")
TITLE_NAT     = Font(bold=True, size=14, color="1a3c6e")
TITLE_CORP    = Font(bold=True, size=14, color="0d5c3a")
SUBHDR_FONT   = Font(bold=True, size=10)
thin          = Side(style="thin", color="d0d7e4")
BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def write_header(ws, row, cols, fill):
    for ci, title in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=title)
        c.fill = fill; c.font = HDR_FONT; c.border = BORDER; c.alignment = CENTER

def write_row(ws, r, vals, fill, bold=False, center_cols=()):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = fill; c.border = BORDER
        c.font = Font(size=9, bold=bold)
        c.alignment = CENTER if ci in center_cols else LEFT

# ── Hoja 1: Clientes Naturales ────────────────────────────────────────────────
wb     = openpyxl.Workbook()
ws_nat = wb.active
ws_nat.title        = "Clientes Naturales"
ws_nat.freeze_panes = "A5"

ws_nat.merge_cells("A1:K1")
ws_nat["A1"].value     = "BancoBanQuito — Catálogo de Clientes Naturales"
ws_nat["A1"].font      = TITLE_NAT
ws_nat["A1"].alignment = CENTER

ws_nat.merge_cells("A2:K2")
ws_nat["A2"].value     = "500 clientes masivos ordenados por cédula | 20% (primeros 100 en creación) tienen 2 cuentas | Saldo: $1,000.00"
ws_nat["A2"].font      = SUBHDR_FONT
ws_nat["A2"].alignment = CENTER

HDR_NAT = ["#","Cédula","Nombre Completo","Email","Teléfono",
           "N° Cuenta Ahorros","Suc. Ahorros",
           "N° Cuenta Corriente","Suc. Corriente",
           "Saldo","Nota"]
write_header(ws_nat, 4, HDR_NAT, HDR_FILL_NAT)

row_num = 5
for i, nat in enumerate(sorted_naturals, 1):
    has_two = nat["cta_corriente"] is not None
    fill = EVEN_N if i % 2 == 0 else ODD_FILL
    nota = "2 cuentas" if has_two else ""
    vals = [i, nat["cedula"], nat["nombre"], nat["email"], nat["tel"],
            nat["cta_ahorros"], nat["suc_ahorros"],
            nat["cta_corriente"] or "—", nat["suc_corriente"] or "—",
            "$1,000.00", nota]
    write_row(ws_nat, row_num, vals, fill, center_cols=(1, 10))
    row_num += 1

# Separador demo
ws_nat.merge_cells(f"A{row_num}:K{row_num}")
sep = ws_nat.cell(row=row_num, column=1, value="── Clientes Demo (creados en initDemoData, sin cuentas automáticas) ──")
sep.font = Font(bold=True, size=9, color="7B6000")
sep.fill = PatternFill("solid", fgColor="FFF8DC")
sep.alignment = CENTER
row_num += 1

for j, demo in enumerate(sorted(DEMO_NATURALS, key=lambda x: x["cedula"]), 1):
    vals = [f"D{j}", demo["cedula"], demo["nombre"], demo["email"], demo["tel"],
            demo["cta_ahorros"], demo["suc_ahorros"],
            demo["cta_corriente"], demo["suc_corriente"],
            "—", "★ Demo"]
    write_row(ws_nat, row_num, vals, DEMO_FILL, bold=True, center_cols=(1, 10))
    row_num += 1

for col, w in zip("ABCDEFGHIJK", [5, 14, 28, 38, 13, 16, 16, 16, 16, 12, 10]):
    ws_nat.column_dimensions[col].width = w
ws_nat.row_dimensions[1].height = 28
ws_nat.row_dimensions[4].height = 30

# ── Hoja 2: Empresas Corporativas ─────────────────────────────────────────────
ws_corp = wb.create_sheet("Empresas Corporativas")
ws_corp.freeze_panes = "A5"

ws_corp.merge_cells("A1:K1")
ws_corp["A1"].value     = "BancoBanQuito — Catálogo de Clientes Corporativos"
ws_corp["A1"].font      = TITLE_CORP
ws_corp["A1"].alignment = CENTER

ws_corp.merge_cells("A2:K2")
ws_corp["A2"].value     = "50 empresas masivas | 3 cuentas c/u: CTE=Operativa, NOM=Nómina, AHO=Impuestos/Reservas | Saldo: $1,000.00"
ws_corp["A2"].font      = SUBHDR_FONT
ws_corp["A2"].alignment = CENTER

HDR_CORP = ["#","RUC","Razón Social","Email",
            "N° Cuenta Operativa","Suc. Operativa",
            "N° Cuenta Nómina","Suc. Nómina",
            "N° Cuenta Impuestos","Suc. Impuestos",
            "Nota"]
write_header(ws_corp, 4, HDR_CORP, HDR_FILL_CORP)

row_num = 5
for i, comp in enumerate(sorted_companies, 1):
    fill = EVEN_C if i % 2 == 0 else ODD_FILL
    vals = [i, comp["ruc"], comp["nombre"], comp["email"],
            comp["cta_operativa"], comp["suc_operativa"],
            comp["cta_nomina"],    comp["suc_nomina"],
            comp["cta_impuestos"], comp["suc_impuestos"],
            "Pagos Masivos"]
    write_row(ws_corp, row_num, vals, fill, center_cols=(1,))
    row_num += 1

ws_corp.merge_cells(f"A{row_num}:K{row_num}")
sep2 = ws_corp.cell(row=row_num, column=1, value="── Empresas Demo (sin cuentas automáticas) ──")
sep2.font = Font(bold=True, size=9, color="7B6000")
sep2.fill = PatternFill("solid", fgColor="FFF8DC")
sep2.alignment = CENTER
row_num += 1

for j, demo in enumerate(sorted(DEMO_COMPANIES, key=lambda x: x["ruc"]), 1):
    vals = [f"D{j}", demo["ruc"], demo["nombre"], demo["email"],
            demo["cta_operativa"], demo["suc_operativa"],
            demo["cta_nomina"],    demo["suc_nomina"],
            demo["cta_impuestos"], demo["suc_impuestos"],
            "★ Demo"]
    write_row(ws_corp, row_num, vals, DEMO_FILL, bold=True, center_cols=(1,))
    row_num += 1

for col, w in zip("ABCDEFGHIJK", [5, 16, 42, 34, 18, 16, 18, 16, 18, 16, 13]):
    ws_corp.column_dimensions[col].width = w
ws_corp.row_dimensions[1].height = 28
ws_corp.row_dimensions[4].height = 30

# ── Hoja 3: Resumen ───────────────────────────────────────────────────────────
ws_res = wb.create_sheet("Resumen")
resumen = [
    ("BancoBanQuito — Resumen de Volumetría", ""),
    ("", ""),
    ("Categoría", "Valor"),
    ("Clientes Naturales masivos", 500),
    ("  - Con 1 cuenta Ahorros (80%, posiciones 101-500)", 400),
    ("  - Con 2 cuentas: Ahorros + Corriente (20%, posiciones 1-100)", 100),
    ("Clientes Naturales demo (sin cuentas automáticas)", 8),
    ("", ""),
    ("Clientes Corporativos masivos", 50),
    ("  - Cuenta Operativa (CTE)", 50),
    ("  - Cuenta Nómina (NOM)", 50),
    ("  - Cuenta Impuestos/Reservas (AHO)", 50),
    ("Empresas Demo (sin cuentas automáticas)", 3),
    ("", ""),
    ("Total cuentas naturales masivas", 600),
    ("Total cuentas corporativas masivas (3 × 50)", 150),
    ("Cuentas adicionales para llegar a 1,500", 750),
    ("TOTAL CUENTAS ACTIVAS EN BD", 1500),
    ("", ""),
    ("Sucursales", "Código"),
    ("Sucursal Norte",   "001"),
    ("Sucursal Sur",     "002"),
    ("Sucursal Centro",  "003"),
    ("Sucursal Valles",  "004"),
    ("Sucursal Digital", "005"),
    ("", ""),
    ("Formato número de cuenta", "SUCURSAL(3) + SECUENCIAL(7) = 10 dígitos"),
    ("Saldo inicial por cuenta", "$1,000.00 USD"),
]
bold_rows = {1, 3, 15, 16, 17, 18, 20}
for ri, (k, v) in enumerate(resumen, 1):
    ws_res.cell(row=ri, column=1, value=k).font = Font(
        bold=(ri in bold_rows), size=10 if ri > 1 else 14,
        color="1a3c6e" if ri == 1 else "000000")
    ws_res.cell(row=ri, column=2, value=v).font = Font(size=10)
ws_res.column_dimensions["A"].width = 55
ws_res.column_dimensions["B"].width = 22

# ── Guardar ───────────────────────────────────────────────────────────────────
out = "/Users/anahy/Desktop/Arquitectura/grupo1_BancoBanQuito/BancoBanQuito_Clientes_Cuentas.xlsx"
wb.save(out)
print(f"Excel generado: {out}")
print(f"  Naturales masivos : {len(naturals_in_order)}")
print(f"  Empresas masivas  : {len(companies_in_order)}")
print(f"  Secuencia final   : {seq - 1}")

# Verificaciones clave
paola = next((n for n in naturals_in_order if n["cedula"] == "0210045704"), None)
if paola:
    print(f"\n  Verificacion Paola Paredes Vargas:")
    print(f"    naturalIndex    : {paola['natural_index']}")
    i = naturals_in_order.index(paola)
    print(f"    Posicion (0-based): {i}  | {i} % 5 = {i % 5} → {BRANCHES[i % 5]['name']}")
    print(f"    Cuenta Ahorros  : {paola['cta_ahorros']}  (esperado: 0020000457)")
